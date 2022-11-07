from tkinter import *
from tkinter import filedialog
from os import path
from pathlib import Path
from treebase import Parser
from database.database import show_database
from openpyxl import Workbook
from shutil import copyfile


#################################################################
#############  ##############
#################################################################
class Setting_database:
    def __init__(self, root):
        ###
        self.database_directory = Parser().get_parser('database', 'directory')
        self.database_directory_copy = Parser().get_parser('database', 'directory_copy')

        ###
        window = Frame(root)
        window.pack(padx=10, pady=10)

        ###
        database_LabelFrame = LabelFrame(window, text='Ustawienia bazy danych')
        database_LabelFrame.grid(row=0, column=0, padx=0, pady=10)

        execl_LabelFrame = LabelFrame(window, text='Zapis pliku')
        execl_LabelFrame.grid(row=1, column=0, padx=0, pady=10)

        ###
        info_excel = '''
Plik exel zostanie zapisany w wybranym folderze do zapisu bazy danych. 
Jeśli plik nie istnieje w wybranym folderze, spróbuj wyszukać w explorerze 
plik o nazwie "database.xlsx". Jeśli nadal nie będzie można go znaleźć, 
skontaktuj się z twórcą programu.
        '''

        ###
        link_database_label = Label(database_LabelFrame, text='Folder zapisu pliku', justify=LEFT, anchor='w', width=25)
        link_database_button = Button(database_LabelFrame, text='Wybierz', width=10, command=self.open_dir_to_database) 
        self.link_database_entry = Entry(database_LabelFrame, width=40) 
        link_database_copy_label = Label(database_LabelFrame, text='Folder zapisu pliku zapasowego', justify=LEFT, anchor='w', width=25)
        link_database_copy_button = Button(database_LabelFrame, text="Wybierz", width=10, command=self.open_dir_to_database_copy)
        self.link_database_copy_entry = Entry(database_LabelFrame, width=40)
        create_copy_database_label = Label(execl_LabelFrame, text='Zapis kopii bazy danych', justify=LEFT, anchor='w', width=25)
        create_copy_database_button = Button(execl_LabelFrame, text='Stwórz', width=10, command=self.copy_database)     
        database_to_exel_label = Label(execl_LabelFrame, text='Zapis bazy danych do pliku exel', justify=LEFT, anchor='w', width=25)
        database_to_exel_button = Button(execl_LabelFrame, text='Stwórz', width=10, command=self.save_to_excel)
        info = Label(execl_LabelFrame, text=info_excel, width=60, justify=LEFT)

        ###
        self.link_database_entry.insert(0, self.database_directory)
        self.link_database_copy_entry.insert(0, self.database_directory_copy)

        ###
        link_database_label.grid(row=0, column=0, padx=(5, 140), pady=5)
        link_database_button.grid(row=0, column=1, padx=10, pady=5)
        self.link_database_entry.grid(row=1, column=0, columnspan=2, padx=(5, 140), pady=(5, 15))
        link_database_copy_label.grid(row=2, column=0, padx=(5, 140), pady=(10, 5))        
        link_database_copy_button.grid(row=2, column=1, padx=10, pady=5)        
        self.link_database_copy_entry.grid(row=3, column=0, columnspan=2, padx=(5, 140), pady=(5,15))
        create_copy_database_label.grid(row=4, column=0, padx=(5, 140), pady=(5, 0))
        create_copy_database_button.grid(row=4, column=1, padx=10, pady=(5, 0))
        database_to_exel_label.grid(row=6, column=0, padx=(5, 140), pady=(5, 5))
        database_to_exel_button.grid(row=6, column=1, padx=10, pady=(5, 5))
        info.grid(row=7, column=0, columnspan=2, padx=0, pady=(5, 15))

    #################################################################
    #############  ##############
    #################################################################
    def open_dir_to_database(self):
        self.link_database_entry.delete(0, END)
        self.link_database_entry.insert(0, self.open_dir('directory'))


    #################################################################
    #############  ##############
    #################################################################
    def open_dir_to_database_copy(self):
        self.link_database_copy_entry.delete(0, END)
        self.link_database_copy_entry.insert(0, self.open_dir('directory_copy'))


    #################################################################
    #############  ##############
    #################################################################
    def open_dir(self, name):
        database_directory = filedialog.askdirectory(initialdir='./')
        file_database = str(Parser().get_parser('database', 'directory') + '/database.db')

        if database_directory:
            Parser().edit_parser("database", name, database_directory)

        new_file_database = str(Parser().get_parser('database', 'directory') + '/database.db')
        if not Path(str(new_file_database)).is_file():
            copyfile(file_database, new_file_database)
            Path(file_database).unlink()

        return database_directory

    #################################################################
    #############  ##############
    #################################################################
    def copy_database(self):
        file_database = str(Parser().get_parser('database', 'directory')) + '/database.db'
        file_database_copy = str(Parser().get_parser('database', 'directory_copy')) + '/database_copy.db'
        copyfile(file_database, file_database_copy)

    #################################################################
    #############  ##############
    #################################################################
    def save_to_excel(self):
        directory = path.join(path.expanduser('~'), 'Dokumenty\\Biblioteka ZHP Pabianice\\')
        directory = self.link_database_entry.get()
        Path(directory).mkdir(parents=True, exist_ok=True)
        file_excel = path.join(directory, 'database.xlsx')

        ###
        workbook = Workbook()
        ###
        list_book = workbook.create_sheet("Lista Książek")
        list_person = workbook.create_sheet("Lista wypożyczających")

        ###
        for row in show_database('listofbook'):
            list_book.append(row)
        
        for row in show_database('personrental'):
            list_person.append(row)

        ###
        remove_sheet = workbook["Sheet"]
        workbook.remove(remove_sheet)

        ###
        workbook.save(file_excel)
        

if __name__ == '__main__':
    Setting_database()